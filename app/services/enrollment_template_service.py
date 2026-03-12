"""Service for generating and parsing Excel enrollment templates."""

import logging
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

# Column definitions: (field_name, human_label, width, required)
ENROLLMENT_COLUMNS = [
    ("first_name", "First Name *", 20, True),
    ("last_name", "Last Name *", 20, True),
    ("date_of_birth", "Date of Birth (YYYY-MM-DD)", 18, False),
    ("gender", "Gender", 12, False),
    ("class_name", "Class Name", 22, False),
    ("medical_info", "Medical Info", 25, False),
    ("allergies", "Allergies", 25, False),
    ("parent_email", "Parent Email", 28, False),
    ("parent_phone", "Parent Phone", 18, False),
    ("emergency_contact_name", "Emergency Contact Name", 22, False),
    ("emergency_contact_phone", "Emergency Contact Phone", 22, False),
]

# Navy brand color
HEADER_BG_COLOR = "1B3A6B"
HEADER_FONT_COLOR = "FFFFFF"
INSTRUCTIONS_FONT_COLOR = "6B7280"


class EnrollmentTemplateService:
    """Generate downloadable Excel templates and parse uploaded ones."""

    @staticmethod
    def generate_template(
        class_names: list[str],
        prefill_class_name: str | None = None,
    ) -> bytes:
        """
        Generate an Excel enrollment template.

        Args:
            class_names: List of class names for dropdown validation.
            prefill_class_name: If provided, pre-fill the class_name column.

        Returns:
            Excel file bytes ready for streaming response.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Enrollment"

        # Styles
        header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
        header_fill = PatternFill(
            start_color=HEADER_BG_COLOR,
            end_color=HEADER_BG_COLOR,
            fill_type="solid",
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        instruction_font = Font(italic=True, color=INSTRUCTIONS_FONT_COLOR, size=10)

        # Row 1: Field name headers (these are the machine-readable keys)
        for col_idx, (field_name, _, width, _) in enumerate(ENROLLMENT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Row 2: Human-readable labels / instructions
        for col_idx, (_, label, _, _) in enumerate(ENROLLMENT_COLUMNS, 1):
            cell = ws.cell(row=2, column=col_idx, value=label)
            cell.font = instruction_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Freeze top 2 rows
        ws.freeze_panes = "A3"

        # Data validation: Gender dropdown (column D, rows 3-500)
        gender_dv = DataValidation(
            type="list",
            formula1='"MALE,FEMALE,OTHER"',
            allow_blank=True,
        )
        gender_dv.error = "Please select MALE, FEMALE, or OTHER"
        gender_dv.errorTitle = "Invalid Gender"
        gender_dv.prompt = "Select gender"
        gender_dv.promptTitle = "Gender"
        ws.add_data_validation(gender_dv)
        gender_dv.add(f"D3:D500")

        # Data validation: Class name dropdown (column E, rows 3-500)
        if class_names:
            # Excel data validation list has a ~255 char limit for formula1
            # Use a comma-separated list within quotes
            class_list_str = ",".join(class_names)
            if len(class_list_str) <= 250:
                class_dv = DataValidation(
                    type="list",
                    formula1=f'"{class_list_str}"',
                    allow_blank=True,
                )
            else:
                # Too many classes for inline list — write to a hidden sheet
                hidden_ws = wb.create_sheet("_classes")
                for i, name in enumerate(class_names, 1):
                    hidden_ws.cell(row=i, column=1, value=name)
                hidden_ws.sheet_state = "hidden"
                class_dv = DataValidation(
                    type="list",
                    formula1=f"=_classes!$A$1:$A${len(class_names)}",
                    allow_blank=True,
                )
            class_dv.error = "Please select a valid class name"
            class_dv.errorTitle = "Invalid Class"
            class_dv.prompt = "Select class"
            class_dv.promptTitle = "Class Name"
            ws.add_data_validation(class_dv)
            class_dv.add(f"E3:E500")

        # Pre-fill class name if specified
        if prefill_class_name:
            for row in range(3, 103):  # Pre-fill 100 rows
                ws.cell(row=row, column=5, value=prefill_class_name)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def parse_excel(file_bytes: bytes) -> tuple[list[dict], int]:
        """
        Parse an uploaded Excel enrollment file.

        Args:
            file_bytes: Raw bytes of the uploaded .xlsx file.

        Returns:
            Tuple of (list of row dicts, total non-empty row count).
            Each dict has keys matching ENROLLMENT_COLUMNS field names.
        """
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        # Field names in column order
        field_names = [col[0] for col in ENROLLMENT_COLUMNS]

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if row is None:
                continue

            # Check if row has actual student data (at least first_name or last_name)
            # Skip rows that only have pre-filled class names but no student info
            if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
                continue

            # first_name is column A (index 0), last_name is column B (index 1)
            first_name_val = row[0] if len(row) > 0 else None
            last_name_val = row[1] if len(row) > 1 else None
            has_first = first_name_val is not None and (not isinstance(first_name_val, str) or first_name_val.strip() != "")
            has_last = last_name_val is not None and (not isinstance(last_name_val, str) or last_name_val.strip() != "")
            if not has_first and not has_last:
                continue

            row_dict = {}
            for col_idx, field_name in enumerate(field_names):
                value = row[col_idx] if col_idx < len(row) else None

                # Convert dates to string format
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%Y-%m-%d")
                elif value is None:
                    value = ""
                else:
                    value = str(value).strip()

                row_dict[field_name] = value

            # Store the original Excel row number for error reporting
            row_dict["_row_number"] = row_idx
            rows.append(row_dict)

        wb.close()
        return rows, len(rows)


def get_enrollment_template_service() -> EnrollmentTemplateService:
    """Get an instance of the enrollment template service."""
    return EnrollmentTemplateService()
