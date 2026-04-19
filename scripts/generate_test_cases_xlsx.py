"""Generate a comprehensive test cases Excel sheet for the ClassUp QA tester."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = "ClassUp_Test_Cases_v2.xlsx"

# ---------- Test cases ----------
# (id, module, feature, description, preconditions, steps, expected, priority, type)
TEST_CASES = [
    # ==================== AUTH ====================
    ("AUTH-001", "Authentication", "Login (happy path)",
     "Admin can log in with valid credentials",
     "School admin account exists",
     "1. Go to /login\n2. Enter valid email + password\n3. Click 'Log In'",
     "Redirected to /dashboard with welcome message. Session cookie set.",
     "Critical", "Smoke"),
    ("AUTH-002", "Authentication", "Login (wrong password)",
     "Invalid password rejected",
     "School admin account exists",
     "1. Go to /login\n2. Enter valid email + wrong password\n3. Click 'Log In'",
     "Error message shown. Still on /login. No session cookie.",
     "Critical", "Negative"),
    ("AUTH-003", "Authentication", "Login (non-existent email)",
     "Unknown email rejected",
     "None",
     "1. Go to /login\n2. Enter unregistered email\n3. Click 'Log In'",
     "Generic auth error (must NOT reveal whether email exists).",
     "High", "Security"),
    ("AUTH-004", "Authentication", "Rate limiting",
     "Login brute force is rate limited",
     "None",
     "1. Attempt login with wrong password 11+ times in 1 minute",
     "After ~10 attempts, receive 429 Too Many Requests error.",
     "High", "Security"),
    ("AUTH-005", "Authentication", "Forgot password",
     "User can request password reset email",
     "User account exists",
     "1. Click 'Forgot Password?' on login\n2. Enter email\n3. Submit",
     "Success message shown. Reset email delivered within 1 minute.",
     "High", "Functional"),
    ("AUTH-006", "Authentication", "Reset password",
     "User can reset password with valid token",
     "Password reset email received",
     "1. Click reset link in email\n2. Enter new password (min 8 chars)\n3. Submit",
     "Password updated. User can log in with new password.",
     "High", "Functional"),
    ("AUTH-007", "Authentication", "Reset password (expired)",
     "Expired reset token rejected",
     "Reset email received more than 24h ago",
     "1. Click reset link\n2. Submit new password",
     "Error: token expired or invalid. Password unchanged.",
     "Medium", "Negative"),
    ("AUTH-008", "Authentication", "Logout",
     "User can log out",
     "Logged in",
     "1. Click profile menu\n2. Click 'Log Out'",
     "Redirected to /login. Session cookie cleared. Cannot access /dashboard.",
     "Critical", "Smoke"),
    ("AUTH-009", "Authentication", "Profile update",
     "User updates own profile",
     "Logged in",
     "1. Go to /profile\n2. Update first name / phone / language\n3. Save",
     "Success toast. Changes persisted after reload.",
     "Medium", "Functional"),
    ("AUTH-010", "Authentication", "Change password",
     "User changes own password",
     "Logged in",
     "1. Go to /profile\n2. Enter current + new password\n3. Submit",
     "Success toast. Log out and log in with new password succeeds.",
     "High", "Functional"),

    # ==================== ONBOARDING ====================
    ("ONB-001", "Onboarding", "Wizard triggers on first login",
     "New school admin sees 5-step onboarding wizard",
     "New tenant created, first admin login",
     "1. Log in for the first time",
     "Redirected to /onboarding step 1.",
     "Critical", "Smoke"),
    ("ONB-002", "Onboarding", "Step 1: School info",
     "Admin fills in school details",
     "On onboarding step 1",
     "1. Enter school name, address, timezone\n2. Upload logo\n3. Pick colors\n4. Next",
     "Progress saved. Advances to step 2.",
     "High", "Functional"),
    ("ONB-003", "Onboarding", "Step 2: Education type + features",
     "Admin picks education type + toggles features",
     "On onboarding step 2",
     "1. Select education type (Daycare/Primary/High School/K12)\n2. Toggle feature switches\n3. Next",
     "Features saved. Terminology adjusts (student vs child). Advances to step 3.",
     "High", "Functional"),
    ("ONB-004", "Onboarding", "Step 3: Create classes",
     "Admin creates at least 1 class",
     "On step 3",
     "1. Enter class name and grade level\n2. Add\n3. Repeat for multiple\n4. Next",
     "Classes saved. Cannot advance with 0 classes.",
     "High", "Functional"),
    ("ONB-005", "Onboarding", "Completion",
     "Wizard completes and marks tenant onboarded",
     "On step 5",
     "1. Review summary\n2. Click Finish",
     "Redirected to /dashboard. Wizard does not re-appear on next login.",
     "High", "Functional"),

    # ==================== STUDENTS ====================
    ("STU-001", "Students", "Create student",
     "Admin adds a new student",
     "Logged in as admin, at least 1 class exists",
     "1. Go to /students\n2. Click 'Add Student'\n3. Fill form (name, DOB, class, etc)\n4. Save",
     "Student appears in list. Success toast.",
     "Critical", "Smoke"),
    ("STU-002", "Students", "Edit student",
     "Admin edits student details",
     "Student exists",
     "1. Go to /students/{id}/edit\n2. Change name/class\n3. Save",
     "Changes persisted. Success toast.",
     "High", "Functional"),
    ("STU-003", "Students", "Search student by name",
     "Admin can search students",
     "Multiple students exist",
     "1. Go to /students\n2. Type partial name in search\n3. Click Filter",
     "List filters to matching students. No UUID errors.",
     "High", "Functional"),
    ("STU-004", "Students", "Filter by class",
     "Filter students by class",
     "Multiple classes with students",
     "1. /students\n2. Select a class from dropdown\n3. Filter",
     "Only students in that class shown.",
     "Medium", "Functional"),
    ("STU-005", "Students", "Filter by grade level",
     "Filter students by grade",
     "Grades assigned",
     "1. Select grade level\n2. Filter",
     "Only matching students shown. Empty grade selection does not error.",
     "Medium", "Functional"),
    ("STU-006", "Students", "Student detail view",
     "View full student profile",
     "Student exists",
     "1. Click a student in list",
     "Shows details, parents, attendance summary, quick actions (Create Report, Invite Parent).",
     "High", "Functional"),
    ("STU-007", "Students", "Export CSV",
     "Export student list to CSV",
     "Students exist",
     "1. /students\n2. Click CSV button",
     "CSV downloads with columns: First, Last, Class, Grade, DOB, Gender, Status, Enrolled.",
     "Medium", "Functional"),
    ("STU-008", "Students", "Export PDF",
     "Export student list to PDF",
     "Students exist",
     "1. /students\n2. Click PDF button",
     "PDF downloads with formatted student table and date.",
     "Medium", "Functional"),
    ("STU-009", "Students", "Print view",
     "Print-friendly student list",
     "Students exist",
     "1. /students\n2. Click Print",
     "Browser print dialog. Nav/buttons hidden. Table fits clearly.",
     "Low", "UI"),
    ("STU-010", "Students", "Delete student",
     "Soft-delete a student",
     "Student exists",
     "1. Open student\n2. Delete\n3. Confirm",
     "Student removed from list. Attendance history preserved.",
     "Medium", "Functional"),

    # ==================== CLASSES ====================
    ("CLS-001", "Classes", "Create class",
     "Admin creates a class",
     "Logged in as admin",
     "1. /classes\n2. Add\n3. Name + grade level + capacity\n4. Save",
     "Class appears in list.",
     "Critical", "Smoke"),
    ("CLS-002", "Classes", "Edit class",
     "Admin edits class",
     "Class exists",
     "1. Open class\n2. Edit\n3. Change name/capacity\n4. Save",
     "Changes saved.",
     "High", "Functional"),
    ("CLS-003", "Classes", "Assign teacher to class",
     "Assign primary + secondary teachers",
     "Class and teacher(s) exist",
     "1. Open class\n2. Teachers section\n3. Add teacher\n4. Optional: Set Primary",
     "Teacher appears in class. Primary flag visible.",
     "High", "Functional"),
    ("CLS-004", "Classes", "View students in class",
     "See all students in a class",
     "Class with students",
     "1. Open class\n2. Students tab",
     "Lists all active students in the class.",
     "Medium", "Functional"),
    ("CLS-005", "Classes", "Teacher class selector (navbar)",
     "Teacher with multiple classes can switch",
     "Teacher with 2+ assigned classes",
     "1. Log in as teacher\n2. Use class dropdown in navbar",
     "Selected class persists across pages. Default is primary class.",
     "Medium", "Functional"),
    ("CLS-006", "Classes", "Delete class (with students)",
     "Cannot accidentally wipe student data",
     "Class with students",
     "1. Try to delete class",
     "Warning shown. Students reassigned or deletion blocked.",
     "High", "Negative"),

    # ==================== TEACHERS ====================
    ("TCH-001", "Teachers", "Invite teacher",
     "Admin invites a new teacher",
     "Logged in as admin",
     "1. /teachers\n2. Invite\n3. Enter name + email\n4. Send",
     "Invitation email sent. Appears in pending invitations list.",
     "High", "Functional"),
    ("TCH-002", "Teachers", "Teacher registers via invite",
     "New teacher completes registration",
     "Invitation sent",
     "1. Click invite link\n2. Set password\n3. Submit",
     "Account created. Auto-login. Redirected to dashboard.",
     "Critical", "Smoke"),
    ("TCH-003", "Teachers", "Teacher list",
     "Admin sees all teachers",
     "1+ teachers exist",
     "1. /teachers",
     "List shows all teachers with status + assigned classes.",
     "Medium", "Functional"),
    ("TCH-004", "Teachers", "Resend invitation",
     "Resend pending teacher invite",
     "Pending invitation exists",
     "1. /teachers\n2. Resend on pending row",
     "New email sent. Code may refresh.",
     "Low", "Functional"),
    ("TCH-005", "Teachers", "Deactivate teacher",
     "Admin disables teacher account",
     "Active teacher exists",
     "1. Open teacher\n2. Deactivate",
     "Teacher cannot log in. Data preserved. Class assignments removed or flagged.",
     "Medium", "Permission"),

    # ==================== PARENTS ====================
    ("PAR-001", "Parents", "Invite parent",
     "Invite parent to a specific student",
     "Student exists",
     "1. Open student\n2. Invite Parent\n3. Enter name + email\n4. Send",
     "Invitation email sent. Code generated.",
     "High", "Functional"),
    ("PAR-002", "Parents", "Parent registers via invite",
     "Parent completes registration via email link",
     "Invitation sent",
     "1. Click link (includes ?code=X&email=Y)\n2. Name/email pre-filled, read-only\n3. Set password\n4. Submit",
     "Account created. Linked to student. Auto-login.",
     "Critical", "Smoke"),
    ("PAR-003", "Parents", "Parent sees only own children",
     "Tenant isolation for parents",
     "Parent with 1 child",
     "1. Log in as parent\n2. Try to access another student's ID in URL",
     "403 Forbidden or redirect. Cannot view other students.",
     "Critical", "Security"),
    ("PAR-004", "Parents", "Parent views child's attendance",
     "Read-only attendance view",
     "Parent linked to student with attendance records",
     "1. /dashboard\n2. Navigate to attendance",
     "Sees only own child's attendance history.",
     "High", "Functional"),
    ("PAR-005", "Parents", "Parent cannot modify attendance",
     "Parents have read-only access",
     "Logged in as parent",
     "1. Try POST to attendance API endpoint",
     "403 Forbidden.",
     "Critical", "Security"),

    # ==================== ATTENDANCE ====================
    ("ATT-001", "Attendance", "Mark attendance (daily)",
     "Teacher marks class attendance",
     "Class with students",
     "1. /attendance\n2. Pick class + date\n3. Toggle each student PRESENT/ABSENT/LATE/EXCUSED\n4. Save",
     "Records saved. Check-in time recorded for PRESENT/LATE.",
     "Critical", "Smoke"),
    ("ATT-002", "Attendance", "Bulk mark all present",
     "One-click mark all present",
     "Class with students",
     "1. Select class\n2. Click 'Mark all Present'",
     "All students set to PRESENT.",
     "Medium", "Functional"),
    ("ATT-003", "Attendance", "Attendance history",
     "View student attendance history",
     "Student with records",
     "1. Student detail\n2. Attendance tab",
     "Calendar/list of past days with status.",
     "High", "Functional"),
    ("ATT-004", "Attendance", "Absence notification",
     "Parent notified of absence",
     "Student marked ABSENT, parent linked, email configured",
     "1. Mark student ABSENT\n2. Wait for queue",
     "Parent receives email + in-app notification.",
     "High", "Functional"),
    ("ATT-005", "Attendance", "Edit attendance record",
     "Teacher corrects a record",
     "Existing record",
     "1. Open daily view\n2. Change status\n3. Save",
     "Status updated. Change timestamp updated.",
     "Medium", "Functional"),
    ("ATT-006", "Attendance", "Attendance stats",
     "Admin sees attendance rates per class",
     "Data exists",
     "1. /attendance/stats",
     "Shows rates by class, chronic absenteeism alerts.",
     "Medium", "Functional"),
    ("ATT-007", "Attendance", "Teacher only sees own classes",
     "Teacher permission scoping",
     "Teacher assigned to 1 of 3 classes",
     "1. Log in as teacher\n2. Attempt to mark attendance for unassigned class",
     "Cannot select unassigned class. API returns 403.",
     "Critical", "Security"),
    ("ATT-008", "Attendance", "Duplicate prevention",
     "Cannot create two records for same student+date",
     "Existing record",
     "1. Submit attendance for same student on same day",
     "Updates existing record rather than creating duplicate.",
     "High", "Functional"),

    # ==================== MESSAGING ====================
    ("MSG-001", "Messaging", "School-wide announcement",
     "Admin sends announcement to all parents",
     "Logged in as admin",
     "1. /messages/compose\n2. Type ANNOUNCEMENT\n3. Subject + body\n4. Send",
     "All parents receive in-app + email. Appears in inbox.",
     "High", "Functional"),
    ("MSG-002", "Messaging", "Class announcement",
     "Teacher sends to own class parents",
     "Teacher with assigned class",
     "1. Compose CLASS_ANNOUNCEMENT\n2. Select class\n3. Send",
     "Only parents of that class receive. Others do not.",
     "High", "Functional"),
    ("MSG-003", "Messaging", "Student-specific message",
     "Message to one student's parents",
     "Student with linked parents",
     "1. Compose STUDENT_MESSAGE\n2. Select student\n3. Send",
     "Only that student's parents see it.",
     "High", "Functional"),
    ("MSG-004", "Messaging", "Parent reply",
     "Parent replies to thread",
     "Parent has received a message",
     "1. Open message\n2. Reply\n3. Send",
     "Reply appears in thread. Sender (teacher/admin) notified.",
     "Medium", "Functional"),
    ("MSG-005", "Messaging", "Inbox tabs",
     "Filter by message type",
     "Various messages exist",
     "1. /messages\n2. Tabs: All / Announcements / Photos / Documents",
     "Each tab filters correctly. Unread badges update.",
     "Medium", "UI"),
    ("MSG-006", "Messaging", "Mark as read",
     "Read flag updates",
     "Unread message",
     "1. Open message",
     "Unread count decreases. Message marked read.",
     "Low", "Functional"),
    ("MSG-007", "Messaging", "Photo sharing",
     "Share photos to parents",
     "Class/student selected",
     "1. Compose\n2. CLASS_PHOTO / STUDENT_PHOTO\n3. Attach photos\n4. Send",
     "Photos uploaded. Parents see them in inbox + photo gallery.",
     "Medium", "Functional"),
    ("MSG-008", "Messaging", "Document sharing",
     "Share documents (PDF/Word)",
     "File ready",
     "1. Compose\n2. CLASS_DOCUMENT\n3. Attach PDF (<10MB)\n4. Send",
     "Recipients see document. Download works.",
     "Medium", "Functional"),

    # ==================== FILES ====================
    ("FIL-001", "Files", "Upload photo",
     "Upload a JPG/PNG under size limit",
     "Admin/teacher logged in",
     "1. /photos\n2. Drag/drop JPG (<5MB)",
     "Uploaded. Appears in gallery. Thumbnail generated.",
     "High", "Smoke"),
    ("FIL-002", "Files", "Upload oversized file",
     "Reject file over size limit",
     "File > 10MB",
     "1. Try to upload 15MB file",
     "Error message: file too large. Upload rejected.",
     "High", "Negative"),
    ("FIL-003", "Files", "Upload disallowed type",
     "Reject .exe or other non-media",
     "Has .exe file",
     "1. Try to upload .exe",
     "Error: file type not allowed.",
     "Critical", "Security"),
    ("FIL-004", "Files", "Download file",
     "User can download uploaded file",
     "File exists",
     "1. Click download",
     "File downloads with original name.",
     "Medium", "Functional"),
    ("FIL-005", "Files", "Delete own file",
     "User deletes file they uploaded",
     "Uploaded file",
     "1. Open file\n2. Delete",
     "File soft-deleted. Gone from gallery.",
     "Medium", "Functional"),
    ("FIL-006", "Files", "Parent cannot delete school files",
     "Read-only for parents",
     "Logged in as parent",
     "1. Try DELETE /api/v1/files/{id}",
     "403 Forbidden.",
     "High", "Security"),

    # ==================== REPORTS ====================
    ("RPT-001", "Reports", "Create daily report (daycare)",
     "Teacher fills daily activity report",
     "Daycare tenant, student in class",
     "1. Student detail > Create Report\n2. Select template\n3. Fill sections\n4. Save Draft",
     "Report saved as DRAFT.",
     "High", "Functional"),
    ("RPT-002", "Reports", "Finalize report",
     "Finalize and notify parents",
     "Draft report exists",
     "1. Open draft\n2. Finalize",
     "Status = FINALIZED. Parents receive notification + email.",
     "High", "Functional"),
    ("RPT-003", "Reports", "Report template CRUD",
     "Admin creates custom template",
     "Logged in as admin",
     "1. Reports > Templates > New\n2. Add sections\n3. Save",
     "Template available for report creation.",
     "Medium", "Functional"),
    ("RPT-004", "Reports", "Parent views report",
     "Parent sees child's finalized reports",
     "Parent + finalized report exists",
     "1. Log in as parent\n2. Child > Reports",
     "Sees only own child's reports. DRAFT not visible.",
     "High", "Security"),
    ("RPT-005", "Reports", "Report card (primary school)",
     "Generate report card with academic grades",
     "Primary school tenant, subjects assigned",
     "1. Create report\n2. Template = Report Card\n3. Fill marks\n4. Finalize",
     "Grades auto-calculated using grading system. Printable output.",
     "High", "Functional"),
    ("RPT-006", "Reports", "/reports/new 'new' bug check",
     "Old broken link regression",
     "Logged in",
     "1. Go to /reports/new",
     "Redirects to /reports/create or shows form. NO 422 UUID error.",
     "High", "Regression"),
    ("RPT-007", "Reports", "Narrative section rendering",
     "Long narrative text renders safely",
     "Report with long text + HTML",
     "1. View finalized report",
     "Text escaped (no script execution). Line breaks preserved.",
     "Medium", "Security"),
    ("RPT-008", "Reports", "Report listing filter",
     "Filter reports by class/student/date",
     "Multiple reports exist",
     "1. /reports\n2. Apply filters",
     "Correct subset shown.",
     "Medium", "Functional"),

    # ==================== ACADEMIC ====================
    ("ACA-001", "Academic", "Create subject",
     "Admin adds a subject",
     "Logged in as admin",
     "1. /academic\n2. Subjects > New\n3. Name + code\n4. Save",
     "Subject created, unique code enforced.",
     "High", "Functional"),
    ("ACA-002", "Academic", "Assign subject to class",
     "Link subject to class with optional marks override",
     "Subject + class exist",
     "1. Class page > Subjects\n2. Add subject\n3. Set total marks\n4. Save",
     "Subject shows in class. Effective total marks = override or default.",
     "High", "Functional"),
    ("ACA-003", "Academic", "Setup default subjects",
     "One-click default creation",
     "Empty academic config",
     "1. Click 'Setup Defaults'",
     "9 primary or 10 high school subjects created.",
     "Medium", "Functional"),
    ("ACA-004", "Academic", "Grading system CRUD",
     "Create custom grading scale",
     "Admin logged in",
     "1. Grading Systems > New\n2. Add grade bands (0-49 F, 50-59 D, ...)\n3. Mark default\n4. Save",
     "Grading system saved. Default flag enforced (only one).",
     "Medium", "Functional"),
    ("ACA-005", "Academic", "Bulk assign subjects",
     "Copy subjects from one class to another",
     "Source class has subjects",
     "1. Target class > Bulk Assign\n2. Pick source\n3. Apply",
     "All source subjects now in target.",
     "Low", "Functional"),
    ("ACA-006", "Academic", "Remove subject from class",
     "Unlink a subject",
     "Class has a subject",
     "1. Remove subject from class",
     "Removed. Existing reports unaffected.",
     "Medium", "Functional"),

    # ==================== TIMETABLE ====================
    ("TT-001", "Timetable", "Feature flag gated",
     "Timetable link only shows when enabled",
     "Primary/high school tenant",
     "1. Log in\n2. Look at sidebar",
     "'Timetable' link visible. If disabled in settings, link hidden.",
     "Critical", "Smoke"),
    ("TT-002", "Timetable", "Configure day (periods + breaks)",
     "Admin sets up school day",
     "First access",
     "1. /timetable/config\n2. Add 6 periods + 1 break (with times)\n3. Save",
     "Config saved. Reload shows same values.",
     "Critical", "Functional"),
    ("TT-003", "Timetable", "Create timetable for class",
     "Admin creates a timetable",
     "Class has subjects assigned",
     "1. /timetable\n2. New Timetable\n3. Pick class + name\n4. Create",
     "Redirected to grid view. Empty cells shown.",
     "Critical", "Smoke"),
    ("TT-004", "Timetable", "Auto-generate draft",
     "Generator fills the grid",
     "Empty timetable with ≥3 subjects",
     "1. Click 'Auto-Generate Draft'\n2. Leave hours blank\n3. Generate",
     "All non-break cells filled. Subjects spread across days. Teachers assigned. Success toast.",
     "Critical", "Functional"),
    ("TT-005", "Timetable", "Weekly hours override",
     "Generator respects custom counts",
     "Class with 3 subjects",
     "1. Generate with {Math: 10, English: 15} and third blank\n2. Submit",
     "Math appears 10×, English 15×, third fills remaining slots (~15).",
     "High", "Functional"),
    ("TT-006", "Timetable", "Manual cell edit",
     "Click cell to edit",
     "Generated timetable",
     "1. Click a cell\n2. Change subject/teacher in modal\n3. Save",
     "Cell updates. AJAX save toast. No page reload needed.",
     "High", "Functional"),
    ("TT-007", "Timetable", "Clear cell",
     "Empty a cell",
     "Cell with a lesson",
     "1. Click cell\n2. Clear Cell",
     "Cell empties. Shows '+ add' placeholder.",
     "Medium", "Functional"),
    ("TT-008", "Timetable", "Skip breaks in generator",
     "Generator never places lessons on breaks",
     "Config with break at period 4",
     "1. Auto-generate",
     "Period 4 cells all show '—' (break). No lessons.",
     "High", "Functional"),
    ("TT-009", "Timetable", "Conflict badge",
     "Teacher double-booked shows warning",
     "Same teacher assigned to 2 active timetables at same slot",
     "1. Create 2 timetables\n2. Assign same teacher to MON/P1 in both",
     "Red outline + '⚠ Conflict' label on affected cell.",
     "High", "Functional"),
    ("TT-010", "Timetable", "Teacher 'My Schedule' view",
     "Teacher sees only their lessons",
     "Teacher with generated timetable",
     "1. Log in as teacher\n2. /timetable → redirects to /my-schedule",
     "Read-only grid with only lessons they teach. Correct class names.",
     "High", "Smoke"),
    ("TT-011", "Timetable", "Parent child schedule view",
     "Parent sees child's class timetable",
     "Parent with 1 child in class with timetable",
     "1. Log in as parent\n2. /timetable",
     "Auto-redirect to /timetable/child/{id}. Read-only grid shown.",
     "High", "Smoke"),
    ("TT-012", "Timetable", "Parent with multiple children",
     "Parent can switch between children",
     "Parent with 2+ children",
     "1. /timetable/child/{first}\n2. Use child picker",
     "Dropdown visible. Switch loads other child's timetable.",
     "Medium", "Functional"),
    ("TT-013", "Timetable", "Print timetable",
     "Print button hides UI chrome",
     "Any timetable grid",
     "1. Click Print",
     "Browser print dialog. Nav/buttons hidden. Grid fits on page.",
     "Medium", "UI"),
    ("TT-014", "Timetable", "Regenerate clears old entries",
     "Re-running generator doesn't duplicate",
     "Existing generated timetable",
     "1. Auto-Generate again",
     "Old entries cleared. New entries replace them. Count stays at max slots.",
     "High", "Functional"),
    ("TT-015", "Timetable", "Tenant isolation",
     "Timetable cannot be accessed by other tenant",
     "Two tenants",
     "1. Log in as Tenant A\n2. Try GET /api/v1/timetable/timetables/{tenantB_id}",
     "404 Not Found. No data leak.",
     "Critical", "Security"),

    # ==================== BILLING ====================
    ("BIL-001", "Billing", "Create fee item",
     "Admin adds a recurring fee",
     "Billing enabled",
     "1. /billing/fee-items\n2. New\n3. Name + amount + frequency (MONTHLY)\n4. Save",
     "Fee item appears. Can be used for invoice generation.",
     "High", "Functional"),
    ("BIL-002", "Billing", "Generate invoices in bulk",
     "Create invoices for all students",
     "Fee items + students",
     "1. /billing/invoices/generate\n2. Pick items + period\n3. Generate",
     "Invoices created per student. Numbered INV-YYYY-NNNN.",
     "Critical", "Functional"),
    ("BIL-003", "Billing", "Send invoice email",
     "Email invoice to parent",
     "DRAFT invoice exists",
     "1. Open invoice\n2. Send",
     "Status SENT. Email with PDF attachment received by parent.",
     "High", "Functional"),
    ("BIL-004", "Billing", "Record payment",
     "Record parent payment",
     "Sent invoice",
     "1. Open invoice\n2. Record Payment\n3. Amount, method, date\n4. Save",
     "Balance reduced. Status updates (PARTIALLY_PAID / PAID).",
     "Critical", "Functional"),
    ("BIL-005", "Billing", "Overpayment",
     "Payment > balance rejected or flagged",
     "Invoice with R100 balance",
     "1. Record R200 payment",
     "Warning shown or amount capped at balance.",
     "Medium", "Negative"),
    ("BIL-006", "Billing", "Overdue detection",
     "Invoices past due date marked OVERDUE",
     "Invoice with past due_date, status SENT",
     "1. Visit /billing dashboard",
     "Status changes to OVERDUE. Parent notified.",
     "High", "Functional"),
    ("BIL-007", "Billing", "Arrears / ageing report",
     "View outstanding balances",
     "Unpaid invoices",
     "1. /billing/arrears",
     "Per-student table with Total Due, Paid, Outstanding, Current/30/60/90+ buckets.",
     "High", "Functional"),
    ("BIL-008", "Billing", "Arrears filter by class",
     "Filter arrears to one class",
     "Multiple classes with debts",
     "1. /billing/arrears\n2. Select class dropdown",
     "Only students in chosen class shown. Totals recalculate.",
     "Medium", "Functional"),
    ("BIL-009", "Billing", "Arrears CSV export",
     "Export arrears report",
     "Data exists",
     "1. /billing/arrears\n2. Export CSV",
     "CSV downloads with all columns.",
     "Medium", "Functional"),
    ("BIL-010", "Billing", "Parent views own statement",
     "Parent sees invoices + balance for children",
     "Parent with invoiced child",
     "1. Log in as parent\n2. /billing",
     "Statement with invoices, payments, balance. Cannot see other families.",
     "Critical", "Security"),
    ("BIL-011", "Billing", "Currency display",
     "Uses tenant's configured currency",
     "Currency set to ZAR",
     "1. View any billing page",
     "Amounts prefixed with ZAR.",
     "Low", "UI"),
    ("BIL-012", "Billing", "Cancel invoice",
     "Mark invoice CANCELLED",
     "DRAFT/SENT invoice",
     "1. Open invoice\n2. Cancel",
     "Status CANCELLED. Excluded from outstanding totals.",
     "Medium", "Functional"),

    # ==================== SUBSCRIPTIONS ====================
    ("SUB-001", "Subscriptions", "Super admin creates plan",
     "Create new subscription plan",
     "Logged in as super admin",
     "1. /admin/subscriptions\n2. Create Plan\n3. Fill + toggle features\n4. Save",
     "Plan appears in list. Description auto-generated from features.",
     "High", "Functional"),
    ("SUB-002", "Subscriptions", "Edit plan features",
     "Toggling features updates description",
     "Plan exists",
     "1. Edit plan\n2. Toggle feature checkboxes\n3. Save",
     "Description updates. Plan-locked features sync to tenants on activation.",
     "High", "Functional"),
    ("SUB-003", "Subscriptions", "Tenant selects plan",
     "School admin switches plan",
     "Plans available",
     "1. Log in as school admin\n2. /subscription\n3. Select plan",
     "Plan active. Features enabled/disabled per plan. Confirmation shown.",
     "High", "Functional"),
    ("SUB-004", "Subscriptions", "Legacy tenant auto-trial",
     "Old tenants auto-enrolled in free trial",
     "Tenant with no subscription record",
     "1. Log in",
     "Trial auto-started. Banner shows remaining days. No forced upgrade.",
     "High", "Functional"),
    ("SUB-005", "Subscriptions", "Extend trial (super admin)",
     "Super admin extends a trial",
     "Trialing subscription exists",
     "1. /admin/subscriptions\n2. Extend button\n3. Pick days (7/14/30/90)\n4. Submit",
     "Trial end date extended by N days. Subscription now TRIALING if was suspended.",
     "High", "Functional"),
    ("SUB-006", "Subscriptions", "Paystack payment flow",
     "Start payment via Paystack",
     "Plan selected",
     "1. /subscription\n2. Initialize Payment\n3. Complete on Paystack",
     "Webhook receives charge.success. Subscription ACTIVE.",
     "High", "Functional"),
    ("SUB-007", "Subscriptions", "Plan activation syncs features",
     "Activating a plan auto-enables/disables its features on the tenant",
     "Plan with features {billing: true, whatsapp: false}",
     "1. As tenant admin, select that plan\n2. Go to Settings > Features",
     "Billing toggle is ON and locked. WhatsApp is OFF and locked. Labels say 'Plan-managed'.",
     "Critical", "Functional"),
    ("SUB-008", "Subscriptions", "Delete plan",
     "Super admin can delete a plan",
     "Plan exists with no active subscribers",
     "1. /admin/subscriptions\n2. Edit plan\n3. Delete\n4. Confirm",
     "Plan removed from list. Cannot delete if any tenant is actively subscribed.",
     "Medium", "Functional"),
    ("SUB-009", "Subscriptions", "Switch plans keeps billing cycle",
     "Active tenant switches plan mid-cycle",
     "Tenant on active plan A",
     "1. As school admin, go to /subscription\n2. Switch to plan B\n3. Confirm",
     "Features update to plan B immediately. Current period end date preserved.",
     "High", "Functional"),
    ("SUB-010", "Subscriptions", "Trial banner on dashboard",
     "Trial countdown shown",
     "Tenant in TRIALING status",
     "1. Log in as school admin\n2. View dashboard",
     "Purple banner shows 'X days left in trial' with 'View Plans' and 'Add Payment' buttons.",
     "High", "UI"),
    ("SUB-011", "Subscriptions", "Past due banner",
     "Past-due warning shown",
     "Subscription in PAST_DUE status",
     "1. Log in",
     "Yellow banner on dashboard urging payment.",
     "High", "UI"),
    ("SUB-012", "Subscriptions", "Suspended → reactivate",
     "Suspended tenant can reactivate",
     "Tenant in SUSPENDED status",
     "1. Log in as school admin\n2. Click 'Reactivate' on red banner\n3. Select plan / pay",
     "Subscription restored. App access unblocked.",
     "Critical", "Functional"),
    ("SUB-013", "Subscriptions", "Public plans endpoint",
     "Marketing site can list plans",
     "None",
     "1. GET /api/v1/plans (no auth)",
     "Returns active plans only with name, price, features, limits. No secrets.",
     "Medium", "Functional"),
    ("SUB-014", "Subscriptions", "Paystack webhook signature",
     "Invalid webhook signatures rejected",
     "Paystack webhook endpoint",
     "1. POST /api/v1/subscription/paystack/webhook with bad signature",
     "401/403. Payload ignored.",
     "Critical", "Security"),
    ("SUB-015", "Subscriptions", "Extend trial input validation",
     "Extend trial rejects invalid days",
     "Trialing subscription",
     "1. As super admin, try to extend trial by 0 days or 500 days via API",
     "Error: 'Extension must be between 1 and 365 days'.",
     "Medium", "Negative"),
    ("SUB-016", "Subscriptions", "Plan feature validation",
     "Feature toggles in plan must be valid keys",
     "Creating new plan",
     "1. Try to save plan with unknown feature key via API",
     "Either silently filtered OR validation error. Unknown keys don't break tenants.",
     "Medium", "Negative"),

    # ==================== SETTINGS ====================
    ("SET-001", "Settings", "Update school info",
     "Admin edits school details",
     "Logged in as admin",
     "1. /settings\n2. School info\n3. Update name/address\n4. Save",
     "Changes persist. Navbar logo updates if changed.",
     "Medium", "Functional"),
    ("SET-002", "Settings", "Toggle feature",
     "Enable/disable a feature",
     "Feature not locked by plan",
     "1. /settings/features\n2. Toggle\n3. Save",
     "Feature updated. Sidebar link appears/disappears.",
     "High", "Functional"),
    ("SET-003", "Settings", "Plan-locked feature",
     "Cannot toggle plan-controlled feature",
     "Active plan with locked feature",
     "1. /settings/features",
     "Locked checkbox disabled. Shows 'Plan-managed' label.",
     "High", "Permission"),
    ("SET-004", "Settings", "Change language",
     "Switch UI to Afrikaans",
     "Logged in",
     "1. Profile > Language = Afrikaans\n2. Reload",
     "All translated strings show in Afrikaans.",
     "Low", "UI"),
    ("SET-005", "Settings", "Email config (super admin)",
     "Change email provider",
     "Super admin",
     "1. /admin/email-settings\n2. Switch to Resend\n3. Enter API key\n4. Test",
     "Test email sent successfully.",
     "High", "Functional"),

    # ==================== SECURITY ====================
    ("SEC-001", "Security", "Cross-tenant data access",
     "Tenant isolation on IDs",
     "Two tenants with data",
     "1. Log in as Tenant A admin\n2. Try GET /api/v1/students/{TenantB_student_id}",
     "404 Not Found. No data returned.",
     "Critical", "Security"),
    ("SEC-002", "Security", "Role escalation attempt",
     "Parent cannot hit admin endpoints",
     "Logged in as parent",
     "1. Try POST /api/v1/students (create)",
     "403 Forbidden.",
     "Critical", "Security"),
    ("SEC-003", "Security", "SQL injection on search",
     "Search input is sanitized",
     "Logged in",
     "1. Search students with \"'; DROP TABLE students; --\"",
     "No error. No data corruption. Results empty or filtered.",
     "Critical", "Security"),
    ("SEC-004", "Security", "XSS in student notes",
     "Script tags escaped",
     "Logged in",
     "1. Edit student notes = '<script>alert(1)</script>'\n2. View profile",
     "Text shown literally. No script execution.",
     "Critical", "Security"),
    ("SEC-005", "Security", "Webhook SSRF",
     "Cannot register localhost webhook",
     "Admin logged in",
     "1. /webhooks\n2. New with URL http://localhost:6379",
     "Rejected with error: 'Webhook URL cannot point to internal addresses.'",
     "High", "Security"),
    ("SEC-006", "Security", "Cookie security",
     "Session cookie has HttpOnly + SameSite",
     "Logged in",
     "1. Check browser dev tools > Cookies > access_token",
     "HttpOnly=true, SameSite=Lax. Secure=true in production.",
     "Critical", "Security"),
    ("SEC-007", "Security", "HTTPS redirect (prod)",
     "HTTP redirects to HTTPS",
     "Production environment",
     "1. Visit http://classup.co.za",
     "301/302 redirect to https://. HSTS header set.",
     "Critical", "Security"),
    ("SEC-008", "Security", "Security headers present",
     "Required security headers",
     "Any page load",
     "1. Check response headers",
     "X-Content-Type-Options: nosniff, X-Frame-Options: DENY, X-XSS-Protection set.",
     "High", "Security"),

    # ==================== UI/UX ====================
    ("UI-001", "UI/UX", "Mobile navigation",
     "Bottom tab nav on mobile",
     "Mobile device (<768px)",
     "1. Open on phone or narrow browser",
     "Bottom nav visible. Sidebar hidden. Touch targets ≥ 44px.",
     "High", "UI"),
    ("UI-002", "UI/UX", "Toast notifications",
     "Success/error toasts appear",
     "Any action",
     "1. Perform save action",
     "Toast appears top-right, auto-dismisses. No native alert() dialogs.",
     "Medium", "UI"),
    ("UI-003", "UI/UX", "Loading states",
     "Buttons disable during async",
     "Any async action",
     "1. Click submit button",
     "Button disabled + spinner until complete.",
     "Low", "UI"),
    ("UI-004", "UI/UX", "Empty states",
     "Friendly empty list screens",
     "Fresh tenant",
     "1. Visit /students / /classes / /reports",
     "Illustration + helpful text + CTA button.",
     "Low", "UI"),
    ("UI-005", "UI/UX", "Error page",
     "Friendly 404 page",
     "None",
     "1. Visit /nonexistent-page",
     "Custom 404 page with link back to dashboard.",
     "Medium", "UI"),
]


# ---------- Styles ----------
HEADER_FILL = PatternFill("solid", start_color="7C3AED")  # brand purple
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
BODY_FONT = Font(size=10, name="Arial")
BODY_BOLD = Font(size=10, name="Arial", bold=True)

THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MODULE_COLORS = {
    "Authentication": "EDE9FE",
    "Onboarding": "DBEAFE",
    "Students": "D1FAE5",
    "Classes": "FEF3C7",
    "Teachers": "FEE2E2",
    "Parents": "FCE7F3",
    "Attendance": "E0E7FF",
    "Messaging": "CFFAFE",
    "Files": "F3E8FF",
    "Reports": "FFEDD5",
    "Academic": "ECFCCB",
    "Timetable": "FEF9C3",
    "Billing": "FEE2E2",
    "Subscriptions": "DBEAFE",
    "Settings": "F5F5F4",
    "Security": "FECACA",
    "UI/UX": "E0F2FE",
}

PRIORITY_COLORS = {
    "Critical": "DC2626",
    "High": "EA580C",
    "Medium": "CA8A04",
    "Low": "6B7280",
}


def set_border(cell):
    cell.border = BORDER


def auto_row_height(ws, start_row, end_row):
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = None  # let excel autofit


# ---------- Build workbook ----------
wb = Workbook()

# === Sheet 1: Instructions ===
ws = wb.active
ws.title = "Instructions"

ws["A1"] = "ClassUp QA Test Plan"
ws["A1"].font = Font(bold=True, size=18, name="Arial", color="7C3AED")
ws.merge_cells("A1:F1")

ws["A2"] = "End-to-end test cases covering every feature of ClassUp v2"
ws["A2"].font = Font(italic=True, size=11, name="Arial", color="6B7280")
ws.merge_cells("A2:F2")

instructions = [
    ("", ""),
    ("How to use this sheet", ""),
    ("1.", "Go to the 'Test Cases' tab."),
    ("2.", "Work through each row in order (tests are grouped by module)."),
    ("3.", "Follow the 'Pre-conditions' and 'Test Steps' columns."),
    ("4.", "Compare the app's behaviour with the 'Expected Outcome' column."),
    ("5.", "Set 'Status' to Pass, Fail, Blocked, or Not Run using the dropdown."),
    ("6.", "Fill in 'Actual Result' with what you observed (important for failures)."),
    ("7.", "Add any bug details or screenshot references in 'Notes / Bugs'."),
    ("8.", "The 'Summary' tab auto-calculates pass rates per module."),
    ("", ""),
    ("Priorities", ""),
    ("Critical", "Must work before any launch. Blocks everything."),
    ("High", "Core feature, must work in first release."),
    ("Medium", "Important but not blocking."),
    ("Low", "Nice-to-have / edge case."),
    ("", ""),
    ("Test types", ""),
    ("Smoke", "Core happy path — must always pass."),
    ("Functional", "Feature works as described."),
    ("Negative", "Invalid input is handled gracefully."),
    ("Security", "Permissions, injection, isolation."),
    ("Permission", "Role-based access control."),
    ("Regression", "Previously-found bug should not return."),
    ("UI", "Look, layout, responsive, copy."),
    ("", ""),
    ("Test accounts (ask dev team)", ""),
    ("Super Admin", "russel@rfm.org.za"),
    ("School Admin", "admin@sunshinedaycare.co.za"),
    ("Teacher", "(provided separately)"),
    ("Parent", "(provided separately)"),
    ("", ""),
    ("Tester", "(Name)"),
    ("Test environment", "(URL)"),
    ("Test run date", "(YYYY-MM-DD)"),
    ("Build / commit", "(git SHA)"),
]

row = 3
for label, text in instructions:
    if label and not text:
        ws.cell(row=row, column=1, value=label).font = Font(
            bold=True, size=12, name="Arial", color="111827"
        )
    elif label:
        ws.cell(row=row, column=1, value=label).font = BODY_BOLD
        ws.cell(row=row, column=2, value=text).font = BODY_FONT
    row += 1

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 80


# === Sheet 2: Test Cases ===
tc = wb.create_sheet("Test Cases")

headers = [
    "Test ID",
    "Module",
    "Feature",
    "Test Case Description",
    "Pre-conditions",
    "Test Steps",
    "Expected Outcome",
    "Priority",
    "Type",
    "Status",
    "Actual Result",
    "Notes / Bugs",
]

for col_idx, h in enumerate(headers, start=1):
    cell = tc.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    set_border(cell)

tc.row_dimensions[1].height = 28

for i, tcdata in enumerate(TEST_CASES, start=2):
    tid, module, feature, desc, pre, steps, expected, priority, ttype = tcdata
    values = [tid, module, feature, desc, pre, steps, expected, priority, ttype, "Not Run", "", ""]
    for col_idx, v in enumerate(values, start=1):
        cell = tc.cell(row=i, column=col_idx, value=v)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        set_border(cell)

    # Module color
    module_fill = PatternFill("solid", start_color=MODULE_COLORS.get(module, "FFFFFF"))
    tc.cell(row=i, column=2).fill = module_fill

    # Priority color text
    pri_cell = tc.cell(row=i, column=8)
    pri_cell.font = Font(size=10, name="Arial", bold=True, color=PRIORITY_COLORS.get(priority, "000000"))
    pri_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Test ID bold
    tc.cell(row=i, column=1).font = BODY_BOLD

# Column widths
widths = {
    "A": 12,
    "B": 16,
    "C": 28,
    "D": 36,
    "E": 28,
    "F": 42,
    "G": 42,
    "H": 11,
    "I": 13,
    "J": 13,
    "K": 28,
    "L": 28,
}
for col, w in widths.items():
    tc.column_dimensions[col].width = w

# Freeze header
tc.freeze_panes = "A2"

# Data validation: Status dropdown
status_dv = DataValidation(
    type="list",
    formula1='"Not Run,Pass,Fail,Blocked"',
    allow_blank=True,
    showDropDown=False,
)
tc.add_data_validation(status_dv)
status_dv.add(f"J2:J{1 + len(TEST_CASES)}")

# Conditional fill for Status via manual PatternFill per row not needed; users pick.
# Auto filter
tc.auto_filter.ref = f"A1:L{1 + len(TEST_CASES)}"


# === Sheet 3: Summary (with formulas) ===
summ = wb.create_sheet("Summary")

summ["A1"] = "Test Run Summary"
summ["A1"].font = Font(bold=True, size=16, name="Arial", color="7C3AED")
summ.merge_cells("A1:F1")

summ["A3"] = "Overall"
summ["A3"].font = Font(bold=True, size=12, name="Arial")

overall_headers = ["Total", "Pass", "Fail", "Blocked", "Not Run", "Pass %"]
for col_idx, h in enumerate(overall_headers, start=1):
    c = summ.cell(row=4, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")
    set_border(c)

last_row = 1 + len(TEST_CASES)
summ["A5"] = f"=COUNTA('Test Cases'!A2:A{last_row})"
summ["B5"] = f'=COUNTIF(\'Test Cases\'!J2:J{last_row},"Pass")'
summ["C5"] = f'=COUNTIF(\'Test Cases\'!J2:J{last_row},"Fail")'
summ["D5"] = f'=COUNTIF(\'Test Cases\'!J2:J{last_row},"Blocked")'
summ["E5"] = f'=COUNTIF(\'Test Cases\'!J2:J{last_row},"Not Run")'
summ["F5"] = "=IF(A5=0,0,B5/A5)"
summ["F5"].number_format = "0.0%"

for col in range(1, 7):
    c = summ.cell(row=5, column=col)
    c.font = BODY_BOLD
    c.alignment = Alignment(horizontal="center")
    set_border(c)

# Per-module summary
summ["A7"] = "Per-Module Breakdown"
summ["A7"].font = Font(bold=True, size=12, name="Arial")

per_mod_headers = ["Module", "Total", "Pass", "Fail", "Blocked", "Not Run", "Pass %"]
for col_idx, h in enumerate(per_mod_headers, start=1):
    c = summ.cell(row=8, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")
    set_border(c)

modules_in_order = []
seen = set()
for tcd in TEST_CASES:
    m = tcd[1]
    if m not in seen:
        modules_in_order.append(m)
        seen.add(m)

for i, module in enumerate(modules_in_order):
    r = 9 + i
    summ.cell(row=r, column=1, value=module).font = BODY_BOLD
    summ.cell(row=r, column=1).fill = PatternFill(
        "solid", start_color=MODULE_COLORS.get(module, "FFFFFF")
    )
    summ.cell(row=r, column=2, value=f'=COUNTIF(\'Test Cases\'!B2:B{last_row},"{module}")')
    summ.cell(
        row=r,
        column=3,
        value=f'=COUNTIFS(\'Test Cases\'!B2:B{last_row},"{module}",\'Test Cases\'!J2:J{last_row},"Pass")',
    )
    summ.cell(
        row=r,
        column=4,
        value=f'=COUNTIFS(\'Test Cases\'!B2:B{last_row},"{module}",\'Test Cases\'!J2:J{last_row},"Fail")',
    )
    summ.cell(
        row=r,
        column=5,
        value=f'=COUNTIFS(\'Test Cases\'!B2:B{last_row},"{module}",\'Test Cases\'!J2:J{last_row},"Blocked")',
    )
    summ.cell(
        row=r,
        column=6,
        value=f'=COUNTIFS(\'Test Cases\'!B2:B{last_row},"{module}",\'Test Cases\'!J2:J{last_row},"Not Run")',
    )
    summ.cell(row=r, column=7, value=f'=IF(B{r}=0,0,C{r}/B{r})')
    summ.cell(row=r, column=7).number_format = "0.0%"
    for col in range(1, 8):
        c = summ.cell(row=r, column=col)
        if col > 1:
            c.font = BODY_FONT
        c.alignment = Alignment(horizontal="center" if col > 1 else "left")
        set_border(c)

# Priority breakdown
prio_start = 9 + len(modules_in_order) + 2
summ.cell(row=prio_start, column=1, value="By Priority").font = Font(
    bold=True, size=12, name="Arial"
)

prio_headers = ["Priority", "Total", "Pass", "Fail", "Pass %"]
for col_idx, h in enumerate(prio_headers, start=1):
    c = summ.cell(row=prio_start + 1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")
    set_border(c)

priorities = ["Critical", "High", "Medium", "Low"]
for i, p in enumerate(priorities):
    r = prio_start + 2 + i
    summ.cell(row=r, column=1, value=p).font = Font(
        size=10, name="Arial", bold=True, color=PRIORITY_COLORS[p]
    )
    summ.cell(row=r, column=2, value=f'=COUNTIF(\'Test Cases\'!H2:H{last_row},"{p}")')
    summ.cell(
        row=r,
        column=3,
        value=f'=COUNTIFS(\'Test Cases\'!H2:H{last_row},"{p}",\'Test Cases\'!J2:J{last_row},"Pass")',
    )
    summ.cell(
        row=r,
        column=4,
        value=f'=COUNTIFS(\'Test Cases\'!H2:H{last_row},"{p}",\'Test Cases\'!J2:J{last_row},"Fail")',
    )
    summ.cell(row=r, column=5, value=f'=IF(B{r}=0,0,C{r}/B{r})')
    summ.cell(row=r, column=5).number_format = "0.0%"
    for col in range(1, 6):
        c = summ.cell(row=r, column=col)
        if col > 1:
            c.font = BODY_FONT
            c.alignment = Alignment(horizontal="center")
        set_border(c)

# Column widths for summary
for col, w in {"A": 22, "B": 10, "C": 10, "D": 10, "E": 10, "F": 12, "G": 12}.items():
    summ.column_dimensions[col].width = w


# Save
wb.save(OUTPUT)
print(f"Created {OUTPUT} with {len(TEST_CASES)} test cases across {len(modules_in_order)} modules")
